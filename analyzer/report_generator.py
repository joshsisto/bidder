"""
Generate reports from auction item data
"""

import os
import time
import pandas as pd
from urllib.parse import quote_plus
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

from config import OUTPUT_DIR, AUCTION_URL, OPENROUTER_ENABLED
from utils.logger import setup_logger

# Set up logger
logger = setup_logger("ReportGenerator")

class ReportGenerator:
    """Generate reports from auction item data"""
    
    @staticmethod
    def generate_excel_report(items, output_file="auction_opportunities.xlsx"):
        """Generate Excel report with sorted opportunities"""
        logger.info("Generating Excel report")
        
        try:
            # Create full path for output file
            output_path = os.path.join(OUTPUT_DIR, output_file)
            
            # Create DataFrame
            data = []
            skipped_items = 0
            for item in items:
                # If LLM is enabled and couldn't identify the product, skip this item
                if OPENROUTER_ENABLED and not item.get('skip_for_processing', False):
                    llm_info = item.get('llm_product_info', {})
                    # Skip if product type or brand is unknown/unclear and LLM was used
                    if (llm_info.get('product_type', '') in ['Unknown', ''] or 
                        llm_info.get('brand', '') in ['Unknown', '']) and 'llm_product_info' in item:
                        skipped_items += 1
                        item['skip_for_processing'] = True
                        logger.info(f"Skipping item {item.get('lotNumber', '')}: LLM couldn't identify product")
                        continue
                
                # Calculate profit margin percentage
                current_bid = item.get('current_bid_float', 0) or 0
                market_price = item.get('market_price', 0) or 0
                potential_profit = market_price - current_bid if market_price > 0 else 0
                
                # Calculate profit margin as a percentage
                profit_margin = 0
                if market_price > 0 and current_bid > 0:
                    profit_margin = (potential_profit / current_bid) * 100
                
                # Format the image URLs as a string
                image_urls = ', '.join(item.get('images', []))
                
                # Extract LLM product information if available
                llm_info = item.get('llm_product_info', {})
                llm_product_info = ""
                if llm_info:
                    # Format the LLM product info as a string
                    product_type = llm_info.get('product_type', 'Unknown')
                    brand = llm_info.get('brand', 'Unknown')
                    model = llm_info.get('model', 'Unknown')
                    attributes = llm_info.get('attributes', 'N/A')
                    
                    llm_product_info = f"{brand} {model} ({product_type}) - {attributes}"
                
                # Create direct search links for verification
                search_query = item.get('used_search_query', '')
                google_search_url = f"https://www.google.com/search?q={quote_plus(search_query)}"
                amazon_search_url = f"https://www.amazon.com/s?k={quote_plus(search_query)}"
                
                data.append({
                    'Lot Number': item.get('lotNumber', ''),
                    'Description': item.get('description', ''),
                    'Enhanced Description': item.get('enhanced_description', ''),
                    'LLM Product Info': llm_product_info,
                    'Used Search Query': search_query,
                    'Current Bid': current_bid,
                    'Market Price': market_price,
                    'Potential Profit': potential_profit,
                    'Profit Margin %': profit_margin,
                    'ROI': profit_margin,  # Added as duplicate column with different header
                    'Time Remaining': item.get('timeRemaining', ''),
                    'Item URL': item.get('itemUrl', ''),
                    'Google Search': google_search_url,
                    'Amazon Search': amazon_search_url,
                    'Image URLs': image_urls,
                    'OCR Text': item.get('ocr_text', '')
                })
            
            # Log how many items were skipped
            if skipped_items > 0:
                logger.info(f"Skipped {skipped_items} items due to insufficient LLM product identification")
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Add metadata
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            auction_name = AUCTION_URL.split('/')[-3]  # Extract auction name from URL
            
            # Sort by potential profit (descending)
            df = df.sort_values(by='Potential Profit', ascending=False)
            
            # Format the Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write the main data
                df.to_excel(writer, sheet_name='Opportunities', index=False)
                
                # Create a metadata sheet
                metadata = pd.DataFrame({
                    'Property': ['Generated On', 'Auction URL', 'Total Items', 'Top Profit Item', 'Items Skipped'],
                    'Value': [
                        timestamp, 
                        AUCTION_URL, 
                        len(df),
                        df.iloc[0]['Description'] if not df.empty else 'None',
                        skipped_items
                    ]
                })
                metadata.to_excel(writer, sheet_name='Metadata', index=False)
                
                # Get the workbook and worksheets
                workbook = writer.book
                worksheet = writer.sheets['Opportunities']
                
                # Convert URLs to hyperlinks
                url_columns = ['Item URL', 'Google Search', 'Amazon Search']
                for col_num, column in enumerate(df.columns):
                    if column in url_columns:
                        for row_num, url in enumerate(df[column], start=2):  # start=2 because Excel is 1-indexed and we have headers
                            if url and isinstance(url, str):
                                # Add hyperlink to cell
                                cell = worksheet.cell(row=row_num, column=col_num+1)  # +1 because Excel columns are 1-indexed
                                cell.hyperlink = url
                                cell.style = 'Hyperlink'
                    
                    # Handle Image URLs column (multiple links separated by commas)
                    elif column == 'Image URLs':
                        # Add a new Image Links column if not already there
                        image_links_col = None
                        for i, col in enumerate(df.columns):
                            if col == 'Image Links':
                                image_links_col = i
                                break
                        
                        if image_links_col is None:
                            # Create a new Image Links column
                            image_links_col = len(df.columns)
                            worksheet.cell(row=1, column=image_links_col+1).value = 'Image Links'
                            
                            # Make the header bold
                            from openpyxl.styles import Font
                            worksheet.cell(row=1, column=image_links_col+1).font = Font(bold=True)
                        
                        # Now add links to each row
                        for row_num, urls_str in enumerate(df[column], start=2):
                            if urls_str and isinstance(urls_str, str):
                                # Split by commas
                                urls = urls_str.split(', ')
                                if urls:
                                    # Create clickable links for each image
                                    link_text = ""
                                    for i, url in enumerate(urls, 1):
                                        url = url.strip()
                                        if url:
                                            link_text += f"Image {i} "
                                            cell = worksheet.cell(row=row_num, column=image_links_col+1)
                                            cell.value = link_text.strip()
                                            # Make the cell a hyperlink to the first image
                                            if i == 1:
                                                cell.hyperlink = url
                                                cell.style = 'Hyperlink'
                
                # Define cell styling
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Apply header styling
                for col_num, column in enumerate(df.columns, start=1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Set column widths and formatting
                for col_num, column in enumerate(df.columns, start=1):  # Excel columns start at 1
                    column_letter = worksheet.cell(row=1, column=col_num).column_letter
                    
                    # Set appropriate widths based on content type
                    if column in ['Description', 'Enhanced Description', 'OCR Text']:
                        # Text columns - limit to reasonable width
                        worksheet.column_dimensions[column_letter].width = 40
                    elif column in ['LLM Product Info', 'Used Search Query']:
                        # Medium-width columns
                        worksheet.column_dimensions[column_letter].width = 30
                    elif column in ['Current Bid', 'Market Price', 'Potential Profit']:
                        # Currency columns
                        worksheet.column_dimensions[column_letter].width = 12
                        # Apply currency format
                        for row_num in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                    elif column in ['Profit Margin %', 'ROI']:
                        # Percentage columns
                        worksheet.column_dimensions[column_letter].width = 12
                        # Apply percentage format
                        for row_num in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.number_format = '0.00"%"'
                    else:
                        # Other columns - dynamically set width
                        try:
                            column_width = max(len(str(column)), df[column].astype(str).map(len).max())
                            # Limit max width and apply some padding
                            column_width = min(column_width + 2, 40)
                            worksheet.column_dimensions[column_letter].width = column_width
                        except:
                            # Fallback for new columns or errors
                            worksheet.column_dimensions[column_letter].width = 15
                
                # Apply alternating row colors
                for row_num in range(2, len(df) + 2):
                    if row_num % 2 == 0:
                        for col_num in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
                # Color-code the profit margin cells
                profit_margin_col = None
                for col_num, column in enumerate(df.columns, start=1):
                    if column == 'Profit Margin %':
                        profit_margin_col = col_num
                        break
                
                if profit_margin_col:
                    for row_num in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_num, column=profit_margin_col)
                        value = cell.value
                        if isinstance(value, (int, float)):
                            if value >= 100:  # Over 100% profit
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value >= 50:  # 50-100% profit
                                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                            elif value < 0:    # Negative profit
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Add conditional formatting to highlight high-potential items
                for row_num in range(2, len(df) + 2):
                    try:
                        profit_cell = worksheet.cell(row=row_num, column=profit_margin_col) if profit_margin_col else None
                        profit_value = profit_cell.value if profit_cell else 0
                        
                        # If profit margin is very high (>100%), highlight the entire row
                        if isinstance(profit_value, (int, float)) and profit_value >= 100:
                            for col_num in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=row_num, column=col_num)
                                # Only highlight if not already colored
                                if cell.fill.start_color.rgb == '00000000' or cell.fill.start_color.rgb == 'FFF9F9F9':
                                    cell.font = Font(bold=True)
                    except Exception as e:
                        logger.error(f"Error formatting row {row_num}: {e}")
            
            logger.info(f"Report saved to {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return None
